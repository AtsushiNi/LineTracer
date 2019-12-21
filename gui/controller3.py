# コマンドの送信用gui
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.clock import Clock
from kivy.garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
from kivy.core.window import Window

import openpyxl

import numpy as np
import matplotlib.pyplot as plt
import matplotlib
# Kivy 上で Matplotlib を使うために必要な準備
matplotlib.use('module://kivy.garden.matplotlib.backend_kivy')
import re
import serial
import threading

import analyseModule

# 定数
FRAME_TIME = 10 # 更新間隔(ms)
FRAME_NUM = 200 # 画面に表示するデータ数
ANALYSE_LOG_FILE_NAME = 'logs/sampleLogBy4Sensors.xlsx' # 制御パラメータ算出に使用するログファイル
SAVE_LOG_FILE_NAME = 'logs/log.xlsx' # ログを記録するファイル名
PORT = 'COM3' # Bluetooth: 'COM3'or'COM4', USB: 'COM5'

# Arduinoの変数
sensor_borders = [160, 220, 210, 155] # センサーの閾値
kp = 0.8
ki = 0
kd = 1.2
basic_speed = 80

commands = [] # 送信待ちコマンド
input_value = '' # 入力中のコマンド

# グラフ描画のstart/stop
is_graph_updating = True

# グラフのデータ
light1 = np.zeros(FRAME_NUM)
light2 = np.zeros(FRAME_NUM)
light3 = np.zeros(FRAME_NUM)
light4 = np.zeros(FRAME_NUM)
pos = np.zeros(FRAME_NUM)
time = np.zeros(FRAME_NUM)
case = np.zeros(FRAME_NUM)
u = np.zeros(FRAME_NUM)
rpow = np.zeros(FRAME_NUM)
lpow = np.zeros(FRAME_NUM)
position_theta = np.zeros(FRAME_NUM)
position_x =  np.zeros(FRAME_NUM)
position_y =  np.zeros(FRAME_NUM)

class MainScreen3(BoxLayout):
    # スライダーの初期値
    initial_kp = kp
    initial_ki = ki
    initial_kd = kd
    initial_basic_speed = basic_speed
    
    def handle_change(self, value):
        global input_value
        input_value = value.decode()

    def handle_submit(self, value):
        global commands
        global input_value
        global sensor_borders
        if ',' in input_value:
            # センサーデータの解析
            params11, params12, params21, params22, params31, params32, params41, params42 = analyseModule.analyseData(input_value.split(','), ANALYSE_LOG_FILE_NAME)

            # 係数の送信
            commands.append("A" + str(params11[-1]) + "a")
            commands.append("B" + str(params11[-2]) + "a")
            commands.append("C" + str(params12[-1]) + "a")
            commands.append("D" + str(params12[-2]) + "a")
            commands.append("E" + str(params21[-1]) + "a")
            commands.append("F" + str(params21[-2]) + "a")
            commands.append("G" + str(params22[-1]) + "a")
            commands.append("H" + str(params22[-2]) + "a")
            commands.append("I" + str(params31[-1]) + "a")
            commands.append("J" + str(params31[-2]) + "a")
            commands.append("K" + str(params32[-1]) + "a")
            commands.append("L" + str(params32[-2]) + "a")
            commands.append("M" + str(params41[-1]) + "a")
            commands.append("N" + str(params41[-2]) + "a")
            commands.append("O" + str(params42[-1]) + "a")
            commands.append("P" + str(params42[-2]) + "a")

            # センサの閾値の送信
            borders = input_value.split(',')
            sensor_borders[0] = borders[0] + sensor_borders[0]
            sensor_borders[1] = borders[1] + sensor_borders[1]
            sensor_borders[2] = borders[2] + sensor_borders[2]
            sensor_borders[3] = borders[3] + sensor_borders[3]

            commands.append("W" + str(sensor_borders[0]) + "a")
            commands.append("X" + str(sensor_borders[1]) + "a")
            commands.append("Y" + str(sensor_borders[2]) + "a")
            commands.append("Z" + str(sensor_borders[3]) + "a")

        # sensor_bordersの変更
        elif input_value[0] is "W":
            sensor_borders[0] = int(input_value[1:-1]) + sensor_borders[0]
            commands.append("W" + str(sensor_borders[0]) + "a")
        elif input_value[0] is "X":
            sensor_borders[1] = int(input_value[1:-1]) + sensor_borders[1]
            commands.append("X" + str(sensor_borders[1]) + "a")
        elif input_value[0] is "Y":
            sensor_borders[2] = int(input_value[1:-1]) + sensor_borders[2]
            commands.append("Y" + str(sensor_borders[2]) + "a")
        elif input_value[0] is "Z":
            sensor_borders[3] = int(input_value[1:-1]) + sensor_borders[3]
            commands.append("Z" + str(sensor_borders[3]) + "a")

        # コマンドの送信
        else:
            commands.append(input_value)
            print(input_value)

    def handle_start(self):
        global commands
        global is_graph_updating
        commands.append("b1a")
        is_graph_updating = True
        print("start")

    def handle_stop(self):
        global commands
        global is_graph_updating
        global light1
        global light2
        global light3
        global light4
        global time
        global FRAME_NUM

        commands.append("b0a")
        is_graph_updating = False

        # 最初の200行は0で埋まっているので削除
        while time[0] == 0:
            light1 = np.delete(light1, 0)
            light2 = np.delete(light2, 0)
            light3 = np.delete(light3, 0)
            light4 = np.delete(light4, 0)
            time = np.delete(time, 0)

        # Excelにログ保存
        try:
            wb = openpyxl.load_workbook(SAVE_LOG_FILE_NAME)
            wb.remove(wb["Sheet1"])
            sheet = wb.create_sheet("Sheet1")
            for t in range(np.size(time)):
                sheet.cell(row=t+1,column=1,value=t+1)
                sheet.cell(row=t+1,column=2,value=time[t])
                sheet.cell(row=t+1,column=3,value=light1[t])
                sheet.cell(row=t+1,column=4,value=light2[t])
                sheet.cell(row=t+1,column=5,value=light3[t])
                sheet.cell(row=t+1,column=6,value=light4[t])
            wb.save(SAVE_LOG_FILE_NAME)
        except IndexError:
            print("ログを保存できませんでした")

        # データ初期化
        light1 =light2 = light3 = light4 = pos = time = np.zeros(FRAME_NUM)
        print("stop")

    def handle_mode_test(self):
        global commands
        commands.append("c2a")
        print("testMode")

    def handle_mode_run(self):
        global commands
        commands.append("c0a")
        print("runMode")

    def handle_kp_change(self, value, event):
        if (not event[1].is_mouse_scrolling) and (event[1].grab_current is event[0]):
            global commands
            global kp
            kp = value
            commands.append("j"+str(value)+"a")
            print("KP: "+str(value))

    def handle_ki_change(self, value, event):
        if (not event[1].is_mouse_scrolling) and (event[1].grab_current is event[0]):
            global commands
            global ki
            ki = value
            commands.append("k"+str(value)+"a")
            print("KI: "+str(value))

    def handle_kd_change(self, value, event):
        if (not event[1].is_mouse_scrolling) and (event[1].grab_current is event[0]):
            global commands
            global kd
            kd = value
            commands.append("l"+str(value)+"a")
            print("KD: "+str(value))

    def handle_speed_change(self, value, event):
        if (not event[1].is_mouse_scrolling) and (event[1].grab_current is event[0]):
            global commands
            global basic_speed
            basic_speed = value
            commands.append("e"+str(value)+"a")
            print("Speed: "+str(value))

class GraphView(BoxLayout):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        global FRAME_TIME
        global FRAME_NUM

        # 初期化に用いるデータ
        x = np.linspace(1,FRAME_NUM,FRAME_NUM)
        y = np.zeros(FRAME_NUM)

        # Figure, Axis を保存しておく
        self.fig, self.ax = plt.subplots(4, facecolor="0.1")
        self.ax[0].tick_params(axis='x', colors="0.8")
        self.ax[0].tick_params(axis='y', colors="0.8")
        self.ax[0].set_facecolor((0.4, 0.4, 0.4, 1))
        self.ax[0].grid(axis='y')
        self.ax[0].set_ylim([-100, 160])
        self.ax[1].tick_params(axis='x', colors="0.8")
        self.ax[1].tick_params(axis='y', colors="0.8")
        self.ax[1].set_facecolor((0.4, 0.4, 0.4, 1))
        self.ax[1].set_ylim([-1, 4.2])
        self.ax[2].tick_params(axis='x', colors="0.8")
        self.ax[2].tick_params(axis='y', colors="0.8")
        self.ax[2].set_facecolor((0.4, 0.4, 0.4, 1))
        self.ax[2].set_ylim([-0.5, 200])
        self.ax[3].set_facecolor((0.4, 0.4, 0.4, 1))

        # 最初に描画したときの Line も保存しておく
        self.line11, = self.ax[0].plot(x, y, label="sensor1")
        self.line12, = self.ax[0].plot(x, y, label="sensor2")
        self.line13, = self.ax[0].plot(x, y, label="sensor3")
        self.line14, = self.ax[0].plot(x, y, label="sensor4")

        self.line21, = self.ax[1].plot(x, y, label="position")
        self.line22, = self.ax[1].plot(x, y, label="activeSensorsNum")
        self.line23, = self.ax[1].plot(x, y, label="PIDresponse")

        self.line31, = self.ax[2].plot(x, y, label="rightPower")
        self.line32, = self.ax[2].plot(x, y, label="leftPower")

        self.line41, = self.ax[3].plot([], [])
        # ウィジェットとしてグラフを追加する
        widget = FigureCanvasKivyAgg(self.fig)
        self.add_widget(widget)

        # frame_time秒ごとに表示を更新するタイマーを仕掛ける
        Clock.schedule_interval(self.update_view, FRAME_TIME/1000)

    def update_view(self, *args, **kwargs):

        global light1
        global light2
        global light3
        global light4
        global pos
        global is_graph_updating
        global case
        global rpow
        global lpow
        global u
        global position_theta
        global position_x
        global position_y

        # is_graph_updating == Falseならグラフ更新しない
        if not is_graph_updating:
            return

        # データを更新する
        x = np.linspace(1,FRAME_NUM,FRAME_NUM)
        y11 = sensor_borders[0] - light1[-FRAME_NUM:]
        y12 = sensor_borders[1] - light2[-FRAME_NUM:]
        y13 = sensor_borders[2] - light3[-FRAME_NUM:]
        y14 = sensor_borders[3] - light4[-FRAME_NUM:]

        y21 = pos[-FRAME_NUM:]
        y22 = case[-FRAME_NUM:]
        y23 = u[-FRAME_NUM:]

        y31 = rpow[-FRAME_NUM:]
        y32 = lpow[-FRAME_NUM:]

        x41 = position_x[FRAME_NUM:]
        y41 = position_y[FRAME_NUM:]
        y42 = position_theta[FRAME_NUM:]

        # Line にデータを設定する
        self.line11.set_data(x, y11)
        self.line12.set_data(x, y12)
        self.line13.set_data(x, y13)
        self.line14.set_data(x, y14)
        self.line21.set_data(x, y21)
        self.line22.set_data(x, y22)
        self.line23.set_data(x, y23)
        self.line31.set_data(x, y31)
        self.line32.set_data(x, y32)
        self.line41.set_data(x41, y41)
        # グラフの見栄えを調整する
        self.ax[0].relim()
        self.ax[0].autoscale_view()
        self.ax[0].legend(loc='upper left')
        self.ax[1].relim()
        self.ax[1].autoscale_view()
        self.ax[1].legend(loc='upper left')
        self.ax[2].relim()
        self.ax[2].autoscale_view()
        self.ax[2].legend(loc='upper left')
        # 再描画する
        self.fig.canvas.draw()
        self.fig.canvas.flush_events()

class Controller3App(App):
    def __init__(self, **kwargs):
        super(Controller3App, self).__init__(**kwargs)
        global commands
        # Arduinoに初期値を送信
        # センサーの閾値
        commands.append("W" + str(sensor_borders[0]) + "a")
        commands.append("X" + str(sensor_borders[1]) + "a")
        commands.append("Y" + str(sensor_borders[2]) + "a")
        commands.append("Z" + str(sensor_borders[3]) + "a")
        # pidパラメーター
        commands.append("j" + str(kp) + "a")
        commands.append("k" + str(ki) + "a")
        commands.append("l" + str(kd) + "a")
        # basic_speed
        commands.append("e" + str(basic_speed) + "a")

    def build(self):
        serialClient = SerialClient()
        serialClient.start()
        screen = MainScreen3()
        Window.size = (1000,700)
        Window.top = 50
        return screen

# シリアル通信受信クラス
class SerialClient():
    def start(self):
        serial_thread = threading.Thread(target=self.serial_method, daemon=True)
        serial_thread.start()
    
    # シリアル通信用スレッドの実装部
    def serial_method(self):
        global commands
        print_flag = 0
        ser = serial.Serial(PORT,9600) # シリアル通信

        while True:
            line = ser.readline()
            try:
                line = line.decode().rstrip('\r\n')
                # print_flag += 1
                # if print_flag > 50:
                #     print(line)
                #     print_flag = 0
                receives = re.split(',', line)
                for receive in receives:
                    x = re.split(':', receive)
                    if x[0] == 'light1':
                        global light1
                        light1 = np.append(light1, int(x[1]))
                    elif x[0] == 'light2':
                        global light2
                        light2 = np.append(light2, int(x[1]))
                    elif x[0] == 'light3':
                        global light3
                        light3 = np.append(light3, int(x[1]))
                    elif x[0] == 'light4':
                        global light4
                        light4 = np.append(light4, int(x[1]))
                    elif x[0] == 'pos':
                        global pos
                        pos = np.append(pos, float(x[1]))
                    elif x[0] == 'time':
                        global time
                        time = np.append(time, int(x[1]))
                    elif x[0] == '':
                        do_nothing = 0
                    elif x[0] == 'case':
                        global case
                        case = np.append(case, int(x[1]))
                    elif x[0] == 'u':
                        global u
                        u = np.append(u, float(x[1]))
                    elif x[0] == 'rpow':
                        global rpow
                        rpow = np.append(rpow, float(x[1]))
                    elif x[0] == 'lpow':
                        global lpow
                        lpow = np.append(lpow, float(x[1]))
                    elif x[0] == 'theta':
                        global position_theta
                        position_theta = np.append(position_theta, float(x[1]))
                    elif x[0] == 'x':
                        global position_x
                        position_x = np.append(position_x, float(x[1]))
                    elif x[0] == 'y':
                        global position_y
                        position_y = np.append(position_y, float(x[1]))

                    # else:
                    #     print(x[0])
                    #     print(x[1])
            except UnicodeDecodeError:
                print('ERROR!!!')
                print(line)

            # 送信
            if len(commands) > 0:
                for command in commands:
                    ser.write(command.encode())
                commands = []   

if __name__ == '__main__':
    Controller3App().run()
